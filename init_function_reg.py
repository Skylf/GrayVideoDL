"""
初始化 FunctionNameReg Excel 表格体系的脚本。
为每个代码文件创建独立的函数注册表格。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 定义项目根目录、核心代码目录和表格存放路径
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
CORE_DIR = os.path.join(PROJECT_ROOT, "Files")  # Files 为项目核心文件夹
TABLE_DIR = os.path.join(PROJECT_ROOT, "ProjectDescription")

# 表头列定义：列名 -> 列宽度
HEADERS = [
    ("函数名/方法名", 20),
    ("类型", 12),
    ("参数", 30),
    ("返回值", 20),
    ("函数意义", 40),
    ("位置", 50),
    ("记录/修改时间", 22),
    ("其他", 20),
]

# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def createFunctionRegFile(file_name: str) -> str:
    """
    为指定的代码文件创建一个函数注册 Excel 文件。

    Args:
        file_name (str): 代码文件名（如 main.py）

    Returns:
        str: 创建的 Excel 文件路径
    """
    # 生成 Excel 文件名：将代码文件名的扩展名替换为 _FunctionNameReg.xlsx
    base_name = os.path.splitext(file_name)[0]
    excel_name = f"{base_name}_FunctionNameReg.xlsx"
    excel_path = os.path.join(TABLE_DIR, excel_name)

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = f"{base_name}函数注册表"

    # 写入表头
    for col_idx, (header, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(excel_path)
    print(f"已创建：{excel_path}")
    return excel_path


def main():
    # 确保表格目录存在
    os.makedirs(TABLE_DIR, exist_ok=True)
    print(f"表格目录：{TABLE_DIR}")

    # 扫描 Files 核心文件夹下的所有 .py 文件，为每个文件创建对应的函数注册表
    python_files = []
    if os.path.isdir(CORE_DIR):
        for py_file in os.listdir(CORE_DIR):
            if py_file.endswith(".py"):
                python_files.append(py_file)

    if python_files:
        for py_file in python_files:
            createFunctionRegFile(py_file)
    else:
        # 如果 Files 中没有 .py 文件，创建一个示例模板
        print("Files 核心文件夹中暂未发现代码文件，创建示例模板...")
        createFunctionRegFile("example.py")


if __name__ == "__main__":
    main()

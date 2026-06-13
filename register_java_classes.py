"""
注册 Java 类/方法到 ProjectDescription 下的对应 FunctionNameReg 表格。
为每个 Java 文件创建独立的函数注册表，记录类和方法信息。
"""

import os
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 项目根目录和表格存放路径
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
TABLE_DIR = os.path.join(PROJECT_ROOT, "ProjectDescription")

# 表头定义
HEADERS = [
    ("函数名/方法名", 20),
    ("类型", 12),
    ("参数", 35),
    ("返回值", 20),
    ("函数意义", 45),
    ("位置", 55),
    ("记录/修改时间", 22),
    ("其他", 20),
]

# 样式定义
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 当前时间戳
NOW = datetime.now().strftime("%Y年%m月%d日%H时%M分%S秒")


def createOrUpdateRegFile(relative_path: str, rows: list[dict]) -> str:
    """
    为指定的 Java 文件创建或更新函数注册 Excel 文件。

    Args:
        relative_path (str): Java 文件的相对路径（相对于项目根目录）
        rows (list[dict]): 要写入的行数据列表，每行包含：
            - name: 函数/方法/类名
            - type: 类型（类/方法/函数）
            - params: 参数
            - return_type: 返回值
            - meaning: 函数意义
            - location: 位置
            - other: 其他

    Returns:
        str: Excel 文件的绝对路径
    """
    # 生成 Excel 文件名：将 Java 文件名（不含路径）作为表名
    file_base = os.path.splitext(os.path.basename(relative_path))[0]
    excel_name = f"{file_base}_FunctionNameReg.xlsx"
    excel_path = os.path.join(TABLE_DIR, excel_name)

    # 检查文件是否已存在，存在则加载，不存在则创建
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        target_row = ws.max_row + 1  # 在已有数据后追加
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = f"{file_base}函数注册表"
        # 写入表头
        for col_idx, (header, width) in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            ws.column_dimensions[chr(64 + col_idx)].width = width
        # 冻结首行
        ws.freeze_panes = "A2"
        target_row = 2  # 从第二行开始写入数据

    # 写入数据行
    for row_data in rows:
        values = [
            row_data.get("name", ""),
            row_data.get("type", ""),
            row_data.get("params", ""),
            row_data.get("return_type", ""),
            row_data.get("meaning", ""),
            row_data.get("location", ""),
            NOW,
            row_data.get("other", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=target_row, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER
        target_row += 1

    # 保存文件
    wb.save(excel_path)
    print(f"已更新：{excel_path}（{len(rows)} 条记录）")
    return excel_path


def main():
    # 确保表格目录存在
    os.makedirs(TABLE_DIR, exist_ok=True)

    # ===== 1. MainActivity.java =====
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/MainActivity.java",
        [
            {
                "name": "MainActivity",
                "type": "类",
                "params": "",
                "return_type": "",
                "meaning": "应用主Activity，管理底部导航栏和Fragment切换",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/MainActivity.java)[MainActivity:26]",
                "other": "继承AppCompatActivity，3个Tab页面的容器",
            },
            {
                "name": "onCreate",
                "type": "方法",
                "params": "savedInstanceState<Bundle>",
                "return_type": "void",
                "meaning": "Activity创建回调，初始化界面、边缘显示、底部导航和Fragment",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/MainActivity.java)[MainActivity->onCreate:43]",
                "other": "设置EdgeToEdge，绑定底部导航选择监听器",
            },
            {
                "name": "switchFragment",
                "type": "方法",
                "params": "fragmentClass<Class<T>>, fragmentRef<Fragment>, tag<String>",
                "return_type": "void",
                "meaning": "通用Fragment切换方法，懒加载创建，隐藏/显示切换",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/MainActivity.java)[MainActivity->switchFragment:122]",
                "other": "使用hide/show替代replace，避免重复创建Fragment实例",
            },
        ],
    )

    # ===== 2. HomeFragment.java =====
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java",
        [
            {
                "name": "HomeFragment",
                "type": "类",
                "params": "",
                "return_type": "",
                "meaning": "首页Fragment，提供视频链接输入框、解析按钮、测试环境和日志反馈面板",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment:34]",
                "other": "集成Chaquopy调用Python yt-dlp，附日志面板显示运行反馈",
            },
            {
                "name": "onCreateView",
                "type": "方法",
                "params": "inflater<LayoutInflater>, container<ViewGroup>, savedInstanceState<Bundle>",
                "return_type": "View",
                "meaning": "创建Fragment视图，初始化控件和事件监听器",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->onCreateView:59]",
                "other": "初始化Handler用于跨线程UI更新",
            },
            {
                "name": "initViews",
                "type": "方法",
                "params": "view<View>",
                "return_type": "void",
                "meaning": "初始化所有界面控件，通过findViewByid获取布局中的控件实例",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->initViews:82]",
                "other": "获取输入框、按钮、日志面板等控件引用",
            },
            {
                "name": "setupListeners",
                "type": "方法",
                "params": "",
                "return_type": "void",
                "meaning": "设置按钮点击事件监听器：解析按钮调yt-dlp，测试按钮检测环境",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->setupListeners:103]",
                "other": "后台线程调用Python，避免阻塞UI",
            },
            {
                "name": "callPythonFunction",
                "type": "方法",
                "params": "functionName<String>, parameter<String>",
                "return_type": "String",
                "meaning": "调用Python桥接模块的指定函数，通过Chaquopy API执行Python代码",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->callPythonFunction:167]",
                "other": "首次调用时自动初始化Chaquopy Python环境",
            },
            {
                "name": "appendLog",
                "type": "方法",
                "params": "text<String>",
                "return_type": "void",
                "meaning": "向日志面板追加一条文本，通过Handler切换到主线程更新UI",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->appendLog:215]",
                "other": "支持从后台线程调用，自动滚动到底部",
            },
            {
                "name": "showToast",
                "type": "方法",
                "params": "message<String>",
                "return_type": "void",
                "meaning": "显示简短的Toast提示消息",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->showToast:252]",
                "other": "封装Toast.makeText调用",
            },
            {
                "name": "simplifyCodec",
                "type": "方法",
                "params": "vcodec<String>",
                "return_type": "String",
                "meaning": "简化视频编码名称，将avc1.640028转为h264，hevc转为h265，av01转为av1",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/home/HomeFragment.java)[HomeFragment->simplifyCodec:513]",
                "other": "用于格式行标签显示，unknown编码截取前8字符",
            },
        ],
    )

    # ===== 3. DownloadFragment.java =====
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadFragment.java",
        [
            {
                "name": "DownloadFragment",
                "type": "类",
                "params": "",
                "return_type": "",
                "meaning": "下载列表Fragment，展示所有下载任务的状态和进度",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadFragment.java)[DownloadFragment:22]",
                "other": "继承Fragment，下载Tab页面",
            },
            {
                "name": "onCreateView",
                "type": "方法",
                "params": "inflater<LayoutInflater>, container<ViewGroup>, savedInstanceState<Bundle>",
                "return_type": "View",
                "meaning": "创建Fragment视图，初始化下载列表RecyclerView和空状态提示",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadFragment.java)[DownloadFragment->onCreateView:35]",
                "other": "后续实现下载列表适配器和数据绑定",
            },
        ],
    )

    # ===== 4. SettingsFragment.java =====
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/ui/settings/SettingsFragment.java",
        [
            {
                "name": "SettingsFragment",
                "type": "类",
                "params": "",
                "return_type": "",
                "meaning": "设置Fragment，提供下载路径、合并开关、画质选择等配置",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/settings/SettingsFragment.java)[SettingsFragment:22]",
                "other": "继承Fragment，设置Tab页面",
            },
            {
                "name": "onCreateView",
                "type": "方法",
                "params": "inflater<LayoutInflater>, container<ViewGroup>, savedInstanceState<Bundle>",
                "return_type": "View",
                "meaning": "创建Fragment视图，初始化设置项控件",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/settings/SettingsFragment.java)[SettingsFragment->onCreateView:32]",
                "other": "后续实现SharedPreferences持久化存储",
            },
        ],
    )

    registerDownloadAdapterUpdates()
    registerDownloadTaskUpdates()
    registerVideoInfoModelUpdates()
    print("\n所有函数注册表更新完成！")


def registerDownloadAdapterUpdates():
    """注册 DownloadAdapter 新增的方法"""
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadAdapter.java",
        [
            {
                "name": "setSpeedSizeText",
                "type": "方法",
                "params": "task<DownloadTask>",
                "return_type": "void",
                "meaning": "设置下载速度和已下载/总大小文本显示，速度>0时显示\"⬇ 速度 · 已下载/总大小\"",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadAdapter.java)[DownloadAdapter.ViewHolder->setSpeedSizeText:234]",
                "other": "下载中和暂停状态均显示，数据不足时隐藏",
            },
            {
                "name": "formatSpeedStatic",
                "type": "方法",
                "params": "bytesPerSec<double>",
                "return_type": "String",
                "meaning": "将字节/秒格式化为可读速度文本（如\"2.3 MiB/s\"），静态方法",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadAdapter.java)[DownloadAdapter.ViewHolder->formatSpeedStatic:259]",
                "other": "支持B/s、KiB/s、MiB/s三级单位",
            },
            {
                "name": "formatBytesStatic",
                "type": "方法",
                "params": "bytes<long>",
                "return_type": "String",
                "meaning": "将字节数格式化为可读大小文本（如\"45.2 MB\"），静态方法",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/ui/download/DownloadAdapter.java)[DownloadAdapter.ViewHolder->formatBytesStatic:272]",
                "other": "支持B、KB、MB三级单位",
            },
        ],
    )


def registerDownloadTaskUpdates():
    """注册 DownloadTask 新增的字段和对应 getter/setter"""
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/model/DownloadTask.java",
        [
            {
                "name": "downloadedBytes",
                "type": "字段",
                "params": "long",
                "return_type": "",
                "meaning": "已下载字节数，从进度文件实时更新，持久化保存",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/DownloadTask.java)[DownloadTask:36]",
                "other": "getDownloadedBytes/setDownloadedBytes获取和设置",
            },
            {
                "name": "totalBytesForProgress",
                "type": "字段",
                "params": "long",
                "return_type": "",
                "meaning": "总字节数，从进度文件实时更新，持久化保存",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/DownloadTask.java)[DownloadTask:37]",
                "other": "getTotalBytesForProgress/setTotalBytesForProgress获取和设置",
            },
            {
                "name": "speed",
                "type": "字段",
                "params": "double",
                "return_type": "",
                "meaning": "下载速度（字节/秒），仅运行时使用，不持久化",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/DownloadTask.java)[DownloadTask:38]",
                "other": "getSpeed/setSpeed获取和设置",
            },
        ],
    )


def registerVideoInfoModelUpdates():
    """注册 VideoInfo 新增的方法"""
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java",
        [
            {
                "name": "aggregateByQuality",
                "type": "方法",
                "params": "",
                "return_type": "List<Format>",
                "meaning": "按画质聚合格式列表，每种画质只保留一个最佳格式，用于去重显示",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo->aggregateByQuality:221]",
                "other": "跳过纯音频格式，按360p/720p/1080p分组",
            },
            {
                "name": "pickBetterFormat",
                "type": "方法",
                "params": "a<Format>, b<Format>",
                "return_type": "Format",
                "meaning": "比较两个同画质的格式，按非锁定>锁定、合并>纯视频、文件大>文件小的优先级返回更优格式",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo->pickBetterFormat:258]",
                "other": "静态私有方法，被aggregateByQuality调用",
            },
        ],
    )


def registerVideoInfoModel():
    """为 VideoInfo 数据模型创建函数注册表"""
    createOrUpdateRegFile(
        "Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java",
        [
            {
                "name": "VideoInfo",
                "type": "类",
                "params": "",
                "return_type": "",
                "meaning": "视频信息数据模型，解析yt-dlp返回的JSON数据",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo:19]",
                "other": "包含标题、时长、缩略图、格式列表等字段",
            },
            {
                "name": "Format",
                "type": "类",
                "params": "formatId<String>, ext<String>, resolution<String>, filesize<long>, vcodec<String>, acodec<String>",
                "return_type": "",
                "meaning": "视频格式子类，描述单个可用格式的信息",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo.Format:43]",
                "other": "含isAudioOnly判断、getFilesizeText大小格式化",
            },
            {
                "name": "parseFromJson",
                "type": "方法",
                "params": "jsonString<String>",
                "return_type": "VideoInfo",
                "meaning": "从JSON字符串解析为VideoInfo对象",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo->parseFromJson:115]",
                "other": "静态方法，解析失败返回error状态的对象",
            },
            {
                "name": "getDurationText",
                "type": "方法",
                "params": "",
                "return_type": "String",
                "meaning": "获取格式化的时长文本（X分X秒）",
                "location": "(Files/app/src/main/java/com/example/grayvideodl/model/VideoInfo.java)[VideoInfo->getDurationText:164]",
                "other": "将秒数转换为中文时长描述",
            },
        ],
    )


def registerPythonModule():
    """为 Python 桥接模块创建函数注册表（重构版 v2）"""
    createOrUpdateRegFile(
        "Files/app/src/main/python/ytdlp_bridge.py",
        [
            {
                "name": "ytdlp_bridge",
                "type": "模块",
                "params": "",
                "return_type": "",
                "meaning": "Python桥接模块（重构版v2），通过Chaquopy在Android中运行，提供yt-dlp的封装接口。核心设计：不再自行解析URL或模拟浏览器，所有链接分析由yt-dlp内部extractor完成",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[ytdlp_bridge:1]",
                "other": "重构版移除了所有自定义解析逻辑",
            },
            {
                "name": "_log",
                "type": "函数",
                "params": "message<str>",
                "return_type": "void",
                "meaning": "输出日志到logcat，通过Chaquopy的stdout重定向机制",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[_log:23]",
                "other": "统一使用[ResolveVideoDL]前缀方便过滤",
            },
            {
                "name": "testEnvironment",
                "type": "函数",
                "params": "",
                "return_type": "str",
                "meaning": "测试Python运行环境，返回Python版本和yt-dlp安装状态",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[testEnvironment:29]",
                "other": "返回JSON格式字符串",
            },
            {
                "name": "_get_common_opts",
                "type": "函数",
                "params": "cookie_file<str>",
                "return_type": "dict",
                "meaning": "构建通用yt-dlp选项字典，使用yt-dlp原生参数配置(cookiefile/impersonate/format_sort)，不再设置自定义请求头或平台检测",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[_get_common_opts:54]",
                "other": "impersonate启用浏览器TLS指纹模拟",
            },
            {
                "name": "extractVideoInfo",
                "type": "函数",
                "params": "video_url<str>, cookie_file<str>",
                "return_type": "str",
                "meaning": "提取指定视频链接的信息（重构版），直接调用yt-dlp提取元数据，不再尝试URL变体或回退方案",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[extractVideoInfo:98]",
                "other": "返回JSON格式字符串，保留yt-dlp原始format_note",
            },
            {
                "name": "downloadVideo",
                "type": "函数",
                "params": "video_url<str>, format_id<str>, output_dir<str>, cookie_file<str>",
                "return_type": "str",
                "meaning": "下载指定视频的特定格式到本地目录（无进度回调版本），直接使用用户原始URL",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[downloadVideo:180]",
                "other": "兼容旧调用接口",
            },
            {
                "name": "downloadVideoWithProgress",
                "type": "函数",
                "params": "video_url<str>, format_id<str>, output_dir<str>, cookie_file<str>, progress_file<str>, cancel_flag_file<str>",
                "return_type": "str",
                "meaning": "下载指定视频并支持进度回调和取消操作，通过progress_hook实时写入进度并检测取消标志",
                "location": "(Files/app/src/main/python/ytdlp_bridge.py)[downloadVideoWithProgress:244]",
                "other": "支持暂停/继续操作",
            },
        ],
    )


if __name__ == "__main__":
    main()
    registerVideoInfoModel()
    registerPythonModule()
